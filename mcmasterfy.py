

"""mcmasterfy.py

Reads part numbers from the first column of mcmasterlist.xlsx and
populates column 3 with thread info (or '-'), column 4 with length,
column 5 with the product title/description, and column 6 with the
product link fetched from McMaster-Carr.

Usage:
  python mcmasterfy.py --input mcmasterlist.xlsx --output mcmasterlist.xlsx

Requirements:
  pip install playwright openpyxl tqdm
  python -m playwright install
"""

import time
import re
import argparse
from pathlib import Path
from urllib.parse import quote_plus, urljoin

from openpyxl import load_workbook

from playwright.sync_api import sync_playwright


def open_workbook_with_retry(path: str):
	while True:
		try:
			return load_workbook(path)
		except (PermissionError, OSError):
			input(f'Unable to open "{path}" because it is locked, not found, or open in Excel. Close the file and press Enter to retry...')


def save_workbook_with_retry(workbook, path: str):
	while True:
		try:
			workbook.save(path)
			return
		except (PermissionError, OSError):
			input(f'Unable to save "{path}" because it is open in Excel. Close the file and press Enter to retry...')


def clean_thread(text: str) -> str:
	if not text:
		return '-'
	s = text.replace('×', 'x').replace('X', 'x').replace('”', '"').replace('“', '"')
	s = ' '.join(s.split()).strip('.,;:')

	# Metric thread such as M6 x 1.0, M6-1.0, or M6x1.0
	m_metric = re.search(r'\b(M\s*\d+(?:\.\d+)?)[\s]*[-x×][\s]*(\d+(?:\.\d+)?)\b', s, re.I)
	if m_metric:
		major = m_metric.group(1).upper().replace(' ', '')
		minor = m_metric.group(2)
		return f'{major}x{minor}'

	# Imperial thread such as 1/4"-20, 1/4 - 20, 10-32, #10-32
	m_frac = re.search(r'\b(\d+(?:/\d+)?)(?:\s*\"?)\s*[-x×]\s*(\d+(?:\.\d+)?)\b', s)
	if m_frac:
		major = m_frac.group(1)
		minor = m_frac.group(2)
		if '/' in major or '"' in text:
			return f'{major}"-{minor}'
		return f'{major}-{minor}'

	# Common numeric thread like 10-32 or #10-32
	m_simple = re.search(r'\b#?\d+-\d+\b', s)
	if m_simple:
		return m_simple.group(0)

	# If the string already looks thread-like, return it compacted.
	if re.search(r'\b(?:M\d+|\d+(?:/\d+)?)[\s]*[-x×][\s]*\d+(?:\.\d+)?\b', s, re.I):
		return s

	return '-'


def parse_description_from_title(title: str) -> str:
	# The page title often contains the full name followed by size info and the site name.
	text = title.split('|')[0].strip()
	if ',' in text:
		return text.split(',')[0].strip()
	return text


def extract_thread_from_page(page) -> str:
	try:
		page.wait_for_selector('body', timeout=3000)
	except Exception:
		pass

	html = page.content()

	# Prefer an explicit Thread row if present.
	m2 = re.search(
		r'<tr[^>]*>\s*<td[^>]*>.*?<span[^>]*>\s*Thread(?:\s*Size)?\s*</span>.*?</td>.*?<td[^>]*>(.*?)</td>',
		html,
		re.S | re.I,
	)
	if m2:
		raw = re.sub(r'<[^>]+>', '', m2.group(1)).strip()
		thread = clean_thread(raw)
		if thread != '-':
			return thread

	# Try the rendered DOM for a Thread row.
	try:
		el = page.query_selector('tr:has-text("Thread") td:last-child, tr:has-text("Thread Size") td:last-child')
		if el:
			text = ' '.join(el.inner_text().split())
			thread = clean_thread(text)
			if thread != '-':
				return thread
	except Exception:
		pass

	# Try the page title / header text if it includes thread size.
	# try:
	# 	title_text = page.title() or ''
	# 	if 'Thread' in title_text:
	# 		thread_text = title_text
	# 		m_title = re.search(r'\b(M\s*\d+(?:\.\d+)?)[\s]*[-x×][\s]*(\d+(?:\.\d+)?)\b', thread_text, re.I)
	# 		if m_title:
	# 			return clean_thread(m_title.group(0))
	# 		m_title2 = re.search(r'\b(\d+(?:/\d+)?)(?:\s*\"?)\s*[-x×]\s*(\d+(?:\.\d+)?)\b', thread_text)
	# 		if m_title2:
	# 			return clean_thread(m_title2.group(0))
	# 		m_title3 = re.search(r'\b#?\d+-\d+\b', thread_text)
	# 		if m_title3:
	# 			return clean_thread(m_title3.group(0))
	# except Exception:
	# 	pass

	# Inspect other rendered text blocks with Thread context.
	try:
		for selector in ['h1', 'h2', 'h3', 'span', 'div', 'p', 'label']:
			for el in page.query_selector_all(selector):
				text = el.inner_text() or ''
				if 'Thread' in text:
					thread = clean_thread(text)
					if thread != '-':
						return thread
	except Exception:
		pass

	return '-'


def extract_length_from_page(page) -> str:
	# Look for the specific Length row in the product detail table.
	html = page.content()
	m = re.search(
		r'<tr[^>]*>\s*<td[^>]*>.*?<span>\s*Length\s*</span>.*?</td>.*?<td[^>]*>.*?<p[^>]*>(.*?)</p>',
		html,
		re.S | re.I,
	)
	if m:
		raw = m.group(1)
		raw = re.sub(r'<span[^>]*>(.*?)</span>', r'\1', raw, flags=re.S | re.I)
		raw = re.sub(r'<[^>]+>', '', raw)
		raw = ' '.join(raw.split())
		if raw:
			return raw

	# Try to get the Length text directly from the rendered DOM.
	try:
		el = page.query_selector('tr:has-text("Length") td:last-child p')
		if el:
			text = el.inner_text()
			text = ' '.join(text.split())
			if text:
				return text
	except Exception:
		pass

	return '-'


def find_first_product_link(page) -> str:
	# Try common selectors in search results; return absolute href or None
	selectors = [
		'a[href*="/product/"]',
		'a[href*="/catalog/"]',
		'.search-result a',
		'.result a',
		'a[data-part-number]'
	]
	for sel in selectors:
		try:
			el = page.query_selector(sel)
			if el:
				href = el.get_attribute('href')
				if href:
					return href
		except Exception:
			continue
	return None


def scrape(workbook_path: str, output_path: str, headless: bool = True, delay: float = 1.0, dry_run: bool = False):
	wb = open_workbook_with_retry(workbook_path)
	ws = wb.active

	# Ensure the output header row matches the desired schema.
	ws.cell(1, 1).value = 'Part #'
	ws.cell(1, 2).value = 'Quantity (est.)'
	ws.cell(1, 3).value = 'Thread'
	ws.cell(1, 4).value = 'Length'
	ws.cell(1, 5).value = 'Description'
	ws.cell(1, 6).value = 'Link'

	with sync_playwright() as p:
		browser = p.chromium.launch(headless=headless)
		page = browser.new_page()

		# iterate rows starting at row 2 (assume header in row 1)
		for row in range(2, ws.max_row + 1):
			part = ws.cell(row, 1).value
			if not part:   #PROGRAM MADE BY TONY PASCOE
				continue

			part_str = str(part).strip()
			print(f'Processing row {row}: {part_str}')

			search_url = f'https://www.mcmaster.com/search?q={quote_plus(part_str)}'
			page.goto(search_url, wait_until='networkidle')

			# attempt to find a product link and follow it
			href = find_first_product_link(page)
			if href:
				if href.startswith('http'):
					product_url = href
				else:
					product_url = urljoin('https://www.mcmaster.com', href)
				try:
					page.goto(product_url, wait_until='networkidle')
				except Exception:
					# fallback: continue on current page (maybe search redirected)
					pass
			else:
				# maybe search redirected directly to product page, or try direct product URL
				product_url = None
				if part_str and re.fullmatch(r"[A-Za-z0-9]+", part_str):
					product_url = f'https://www.mcmaster.com/{part_str}/'
					try:
						page.goto(product_url, wait_until='networkidle')
					except Exception:
						product_url = None

			# extract title / description
			try:
				title_text = page.title()
				description = parse_description_from_title(title_text)
			except Exception:
				description = ''

			thread = extract_thread_from_page(page)
			length = extract_length_from_page(page)

			# Capture the final product URL
			final_product_url = page.url

			print(f'  -> description: {description!r}, length: {length!r}, thread: {thread!r}, link: {final_product_url!r}')

			if not dry_run:
				ws.cell(row, 3).value = thread
				ws.cell(row, 4).value = length
				ws.cell(row, 5).value = description
				ws.cell(row, 6).value = final_product_url

			time.sleep(delay)

	if not dry_run:
		save_workbook_with_retry(wb, output_path)
		print(f'Saved results to {output_path}')


def main():
	ap = argparse.ArgumentParser()
	ap.add_argument('--input', '-i', default='mcmasterlist.xlsx')
	ap.add_argument('--output', '-o', default=None)
	ap.add_argument('--headless', action='store_true', default=True)
	ap.add_argument('--no-headless', dest='headless', action='store_false')
	ap.add_argument('--delay', type=float, default=1.0)
	ap.add_argument('--dry-run', action='store_true')
	args = ap.parse_args()

	base_dir = Path(__file__).resolve().parent
	input_path = Path(args.input)
	if not input_path.is_absolute():
		input_path = base_dir / input_path

	output_path = Path(args.output) if args.output else input_path
	if not output_path.is_absolute():
		output_path = base_dir / output_path

	scrape(str(input_path), str(output_path), headless=args.headless, delay=args.delay, dry_run=args.dry_run)


if __name__ == '__main__':
	main()


